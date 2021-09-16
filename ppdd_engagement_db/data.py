from openpyxl import load_workbook, Workbook
from typing import Dict
from datamaps.process import Cleanser
import datetime
from datetime import date
from collections import OrderedDict
from pathlib import Path
import platform


def _platform_docs_dir() -> Path:
    #  Cross plaform file path handling
    if platform.system() == "Linux":
        return Path.home() / "Documents" / "ppdd_engagement_db"
    if platform.system() == "Darwin":
        return Path.home() / "Documents" / "ppdd_engagement_db"
    else:
        return Path.home() / "Documents" / "ppdd_engagement_db"


root_path = _platform_docs_dir()


org_dict = {
    "DfT(c)": "",
    "Highways England": "HE",
    "HS2 Ltd": "HS2",
    "Network Rail": "NR",
    "IPA": "IPA",
    "DEFRA": "DEF",
    "MCA": "MCA",
}


def handle_nones(data):
    if data is None:  # might have to change None to NULL
        return ''
    else:
        return data


def handle_nulls(data):
    if data is None:
        return 'NULL'
    else:
        return data


def calculate_stakeholder_id(single_tuple, id_list):
    # calculate a composite stakeholder ID
    first_initial = single_tuple[1][0]
    second_initial = single_tuple[2][0]
    org = org_dict[single_tuple[4]]
    group = handle_nones(single_tuple[5])
    stakeholder_id = first_initial + second_initial + org + group
    if stakeholder_id not in id_list:
        single_tuple.insert(0, stakeholder_id)
        id_list.append(stakeholder_id)
    else:
        print(stakeholder_id)
        pass

    return id_list


def calculate_entity_id(single_tuple, id_list):
    # calculate a composite stakeholder ID
    type = single_tuple[1][:4].upper()
    abb = single_tuple[3]
    stakeholder_id = abb + type
    if stakeholder_id not in id_list:
        single_tuple.insert(0, stakeholder_id)
        id_list.append(stakeholder_id)
    else:
        print(stakeholder_id)
        pass

    return single_tuple


def get_data(master_file: str, ws_name: str, last_col: int) -> Dict:
    wb = load_workbook(master_file)
    ws = wb[ws_name]
    full_list = []
    id_list = []
    for row in range(2, 29):
        single_tuple = []  # starts life as a list
        for col in range(1, last_col):
            data = ws.cell(row=row, column=col).value
            if isinstance(data, datetime.datetime):
                data = data.date()
            # if col == 2:
            #     if len(data) == 4:
            #         single_tuple.append("ENG" + data[-1])
            #     if len(data) == 5:
            #         single_tuple.append("ENG" + data[-2:])
            #     # single_tuple.append("ENG" + str(data))
            # else:
            single_tuple.append(data)
        # calculate_entity_id(single_tuple, id_list)
        full_list.append(tuple(single_tuple))

    return full_list


def place_data_excel(s_data):
    wb = Workbook()
    ws = wb.active

    for x, t in enumerate(s_data):
        for i, d in enumerate(t):
            ws.cell(row=x+1, column=i+1).value = d

    return wb


# stakeholder_data = ppdd_data(root_path / "ppdd_engagement_db_tables.xlsx", "Stakeholders", 9)
# project_data = ppdd_data(root_path / "ppdd_engagement_db_tables.xlsx", "Projects", 7)
# ppdd_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "PPDDs", 7)
# engagement_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "Engagements", 6)
# ps_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "ProjectStakeholders", 4)
# p_engagements_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "ProjectEngagements", 4)
# engagements_s_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "EngagementStakeholders", 4)
# engagements_ppdd_data = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "EngagementPPDDs", 4)
# engagement_types = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "EngagementTypes", 4)
engagement_ws = get_data(root_path / "ppdd_engagement_db_tables.xlsx", "EngagementWorkStreams", 4)


wb = place_data_excel(engagement_ws)
wb.save(root_path / "ppdd_engage_table.xlsx")